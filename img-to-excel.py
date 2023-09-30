from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image

def img_to_excel(img_path: str, grid_width: int, grid_height: int, output_excel_path:str) -> None:
    """
     Convert a images to an Excel sheet, where the background color of each cell corresponds
     to a pixel's RGB value from the image. User has to provide the grid size of the output excel file.
     :param img_path: The path to the image file.
     :param grid_width: The number of columns in the output excel.
     :param grid_height: The number of rows in the output excel. If 0 is provided the value is calculated
     :param output_excel_path: The path of the output excel file.
    """
    # Open the image
    img = Image.open(img_path)
    
    # If grid_height was not provided, calculate it using the grid_width and the image's aspect ratio
    if not grid_height:
        # Get original image dimensions
        original_width, original_height = img.size
        aspect_ratio = original_height / original_width
        grid_height = int(grid_width * aspect_ratio)

    # Resize the image based on the grid dimensions
    img = img.resize((grid_width, grid_height))
    
    # Create a new Excel workbook and get the active worksheet
    wb = Workbook()
    ws = wb.active
    
    # Define square cell dimensions
    SQUARE_WIDTH = 4  # Adjust this value as needed
    SQUARE_HEIGHT = 18  # Adjust this value as needed
    
    # Set the column width
    for col_num in range(1, grid_width+1):
        ws.column_dimensions[get_column_letter(col_num)].width = SQUARE_WIDTH 
    for row_num in range(1, grid_height+1):
        ws.row_dimensions[row_num].height = SQUARE_HEIGHT

    # Iterate over the pixels of the resized image and set the color to the cells
    for x in range(grid_width):
        for y in range(grid_height):
            pixel = img.getpixel((x, y))
            # Convert the pixel RGB tuple to a fill color for the cell
            fill_color = "{:02X}{:02X}{:02X}".format(*pixel[:3])  # Convert (R, G, B) to HEX format
            cell_fill = PatternFill(fgColor=fill_color, fill_type="solid")
            ws.cell(row=y+1, column=x+1).fill = cell_fill
    
    # Save the Excel workbook to the specified output path
    wb.save(output_excel_path)

if __name__ == '__main__':
    # Example usage
    img_path = input("Enter the path to the PNG image: ")
    grid_width = int(input("Enter the desired grid width: "))
    grid_height = int(input("Enter the desired grid height. 0 to calculate the height based on the width and image aspect ratio: "))
    output_excel_path = input("Enter the path for the output Excel file: ")
    
    img_to_excel(img_path, grid_width, grid_height, output_excel_path)
    print("PNG image has been converted to Excel with cell colors!")